import os

import bpmsim
import openpyxl

number_of_simulations_run = 0


class Environment:
    """ Execution interface for simulation of processes """

    def __init__(self,
                 file_name: str,
                 file_path: str,
                 number_of_activities: int,
                 activity_names: [str],
                 duration_activities: [callable],
                 duration_instance_gen: callable,
                 amount_of_instances_sim: int,
                 amount_needed_for_activity: [int]):
        """ TODO: docstring """
        self.file_name = file_name
        self.file_path = file_path
        self.number_of_activities = number_of_activities
        self.activity_names = activity_names,
        self.duration_activities = duration_activities,
        self.duration_instance_gen = duration_instance_gen
        self.amount_of_instances_sim = amount_of_instances_sim
        self.amount_needed_for_activity = amount_needed_for_activity
        self.simulation_ran = False

    def run(self, resource_allocation: [int]) -> None:
        """ TODO: docstring """
        assert len(resource_allocation) == self.number_of_activities, \
            "Number of resource allocation values is not the same as number of activities in process model"

        process = bpmsim.importBPMN(modelName=self.file_name,
                                    withSubprocess=False,
                                    outputHelp=False,
                                    fileName=self.file_path)

        resources = []
        for i in range(self.number_of_activities):
            current_resource = bpmsim.resource(name="Human" + str(i), capacity=resource_allocation[i])
            resources.append(current_resource)
            bpmsim.setActivity(BPMN=process,
                               name=self.activity_names[0][i],
                               resources=[current_resource],
                               amount=[self.amount_needed_for_activity[i]],
                               duration=self.duration_activities[0][i])

        instance = bpmsim.createInstancesInterDistribution(self.file_name,
                                                           startTime=0,
                                                           time=None,
                                                           number=self.amount_of_instances_sim,
                                                           interval=self.duration_instance_gen)

        sim = bpmsim.simulation(instances=[instance],
                                resources=resources)
        sim.simulation(fileName="./gen_sim/SIMULATION_output")
        sim.simInfo(fileName="./gen_sim/SIMULATION_info")
        sim.simPlot(fileName="./gen_sim/SIMULATION_plot")
        self.simulation_ran = True
        global number_of_simulations_run
        number_of_simulations_run += 1
        print(("\rNumber of bpmsim simulations/fitness func evaluations, with " + str(self.amount_of_instances_sim) +
              " instances per simulation, run: " + str(number_of_simulations_run)), end='')

    def get_average_waiting_times(self) -> [float]:
        """ TODO: docstring """
        assert self.simulation_ran, "Simulation has to be run to get waiting times"
        wb = openpyxl.load_workbook("./gen_sim/SIMULATION_info.xlsx")
        assert "Activities" in wb.sheetnames, \
            "Sheet with info on waiting times ('Activites') not found in bpmsim info Excel"
        ws = wb['Activities']
        home_column = 'D'
        first_row = 4
        average_waiting_times: [float] = []
        for i in range(self.number_of_activities):
            cell_loc = home_column + str((first_row + i))
            average_waiting_times.append(ws[cell_loc].value)
        assert len(average_waiting_times) == self.number_of_activities, \
            "Error while getting the waiting times; Number of waiting times != number of activites"
        return average_waiting_times
